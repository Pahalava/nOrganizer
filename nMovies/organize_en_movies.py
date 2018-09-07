import os
import openpyxl
import organizer
import time
import re

# sel_path = os.getcwd()
sel_path = "D:\\nVideos\\clean videos\\Test"
resource_path = "D:\\nVideos\\clean videos\\Test\\resources"

metadata_type = "nMovies"
metadata_subtype = "Others"

helper = organizer.Organizer(resource_path, sel_path, metadata_type, metadata_subtype)

folders_en = ("_unsorted", "_others", "0-9", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
              "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z")

file_counter, folders, files = 1, 0, 0
ws = None


# prepare folders - Done
# check if the master file exists, if not copy the master file - Done
# scan unsorted folder
# if mkv then clean audio and subtitles to save some space
# Update master file with calculated new paths
# Move files
# Update master file


def init():
    print("Initializing.......")
    helper.create_folders(folders_en)
    helper.validate_master_file()
    start_scan()


def start_scan():
    global file_counter, folders, files, ws
    wb = openpyxl.load_workbook(helper.get_base_metadata(False))
    start_time = time.strftime("%d-%b-%Y %I.%M.%S %p")
    print("start_time:", start_time)
    ws = wb.create_sheet("Scan_" + start_time)
    wb.active = ws
    ws["B1"] = "Path"
    ws["C1"] = helper.get_scan_directory()
    ws["B2"] = "Time"
    ws["C2"] = start_time
    ws["B3"] = "Folders"
    ws["B4"] = "Files"
    ws["B5"] = "Target Folder"
    ws.append([])
    ws.append(["#", "Type", "Name", "File Type", "New Name", "New Path", "Status", "Full path"])
    print()
    print("Scanning...........please wait...")
    print()
    scan_start = time.time()
    scan(helper.get_scan_directory())
    print("Scan completed in %s seconds." % (time.time() - scan_start))
    print("Scanned Folders: ", folders)
    print("Scanned Files:", files)
    ws["C3"] = folders
    ws["C4"] = files
    wb.save(filename=helper.get_base_metadata(False))
    os.startfile(helper.get_base_metadata(False))


def scan(path):
    global file_counter, folders, files, ws
    for entry in os.scandir(path):
        if entry.is_dir():
            ws.append([file_counter, "Folder", entry.name, "", "", "", "SKIP", entry.path])
            folders += 1
            file_counter += 1
            scan(entry.path)
            print("Scanning folder: ", entry.path)
        else:
            ws.append(
                [file_counter, "File", entry.name, os.path.splitext(entry.name)[1][1:], "", identity_folder(entry.name),
                 "TODO", entry.path])
            files += 1
            file_counter += 1
    return


def identity_folder(name):
    pattern =
    print("name:", name)
    folder = helper.get_sub_base_directory()
    first_letter = name[0]
    print("Checking:",first_letter, helper.is_in_char_range(first_letter.upper(), 'A', 'Z'))
    if first_letter.isnumeric():
        folder += "\\" + "0-9"
    elif helper.is_in_char_range(first_letter.upper(), 'A', 'Z'):
        folder += "\\" + first_letter.upper()
    else:
        folder += "\\" + "_others"
    return folder


#init()
print('H' in range(ord('A'), ord('Z') + 1))
