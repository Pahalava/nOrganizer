import os
from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active
count, folders, files = 1,0,0
filename = 'movies.xlsx'
langs = ('தமிழ்/Tamil', 'English/English')
tasks = ('Scan & export', 'Prepare Folders', 'Organize')
ta_fldrs = ('0-9', 'unsorted', 'உயிர்\அ','உயிர்\ஆ','உயிர்\இ','உயிர்\ஈ','உயிர்\உ','உயிர்\ஊ','உயிர்\எ','உயிர்\ஏ','உயிர்\ஐ','உயிர்\ஒ','உயிர்\ஓ','உயிர்\ஔ','மெய்\க','மெய்\ங','மெய்\ச','மெய்\ஞ','மெய்\ட','மெய்\ண','மெய்\த','மெய்\ந','மெய்\ப','மெய்\ம','மெய்\ய','மெய்\ர','மெய்\ல','மெய்\வ','மெய்\ழ','மெய்\ள','மெய்\ற','மெய்\ன')
en_fldrs = ('0-9', 'unsorted', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z')
sel_path = os.getcwd()
sel_lang = 0

def init():
    global sel_path, sel_lang
    print('Curent Working Directory: ', sel_path)
    print()
    sel_lang = choice(langs,'Select language: ')
    sel_task = choice(tasks)
    if(sel_task == 0):
        prepare_export()
    elif(sel_task == 1):
        if (sel_lang == 0):
            prepare(ta_fldrs)
        elif (sel_lang == 1):
            prepare(en_fldrs)
        else:
            print('wrong lang selection')
    elif (sel_task == 2):
        organize()
    else:
        print('wrong task selection')

def choice(atup, msg='Please select an option: '):
    for i, v in enumerate(atup):
        print(i+1,v)
    sel_val = int(input(msg))-1
    print('Selected option: ', atup[sel_val])
    print()
    return sel_val

def prepare(fldrs):
    global sel_path
    os.chdir(sel_path)
    count=0
    for fldr in fldrs:
        try:
            os.makedirs(fldr, exist_ok=True)
            count+=1
        except OSError as e:
            print('Error occured for',fldr,': ', e.args)
    print(count,'of',len(fldrs),'folders created!')

def prepare_export():
    global sel_path, count, folders, files
    ws['B1'] = 'Path'
    ws['C1'] = sel_path
    ws['B2'] = 'Time'
    ws['C2'] = time.strftime('%I:%M%p on %b %d, %Y in %Z')
    ws['B3'] = 'Folders'
    ws['B4'] = 'Files'
    ws['B5'] = 'Target Folder'
    ws.append([])
    ws.append(['#', 'Type', 'Name', 'File Type', 'New Name', 'New Path', 'Status', 'Full path'])
    print()
    print('Scanning...........please wait...')
    print()
    start_time = time.time()
    scan_export(sel_path)
    print('Scan completed in %s seconds.' % (time.time() - start_time))
    print('Scanned Folders: ', folders)
    print('Scanned Files:', files)
    ws['C3'] = folders
    ws['C4'] = files
    wb.save(filename = filename)
    os.startfile(filename)
    
def scan_export(path):
    global count, folders, files
    for entry in os.scandir(path):
        if entry.is_dir():
            ws.append([count, 'Folder', entry.name, '', '', '', 'SKIP', entry.path])
            folders+=1
            count+=1
            scan_export(entry.path)
            print('Scanning folder: ', entry.path)
        else:
            ws.append([count, 'File', entry.name, os.path.splitext(entry.name)[1][1:], '', '', 'TODO', entry.path])
            files+=1
            count+=1
    return

def file_move(path, new_path, new_name):
    os.makedirs(new_path, exist_ok=true)
    shutil.move(path, new_path+'\\'+new_name)
    return 'Done'

def organize():
    print('Coming soon.....')

init()
#input('Press any key to exit')
