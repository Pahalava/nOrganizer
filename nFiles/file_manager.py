import os
import openpyxl

filename = 'move-rename files.xlsx'
tst_folder = 'test'
sel_path = os.getcwd()
count, success, error = 0,0,0
tasks = ('Prepare for a test', 'Test', 'Run', 'Clean', 'Exit')

os.chdir(sel_path)
wb = openpyxl.load_workbook(sel_path+'\\'+filename)
ws = wb.active
start_row = 5
max_row = ws.max_row

def init():
    print('Curent Working Directory: ', sel_path)
    print()
    sel_task = choice(tasks)
    if(sel_task == 0):
        prep_tst()
        save_open()
        init()
    elif(sel_task == 1):
        tst()
        save_open()
        init()
    elif(sel_task == 2):
        run()
        save_open()
        init()
    elif(sel_task == 3):
        clean_fldrs(sel_path)
        init()
    elif(sel_task == 4):
        print('Good bye!')
    else:
        print('wrong lang selection')
        init()

def choice(atup, msg='Please select an option: '):
    for i, v in enumerate(atup):
        print(i+1,v)
    sel_val = int(input(msg))-1
    print('Selected option: ', atup[sel_val])
    print()
    return sel_val

def prep_tst():
    global count, success, error
    os.makedirs(tst_folder, exist_ok=True)
    #os.mkdir(tst_folder)
    os.chdir(tst_folder)
    fname = ''
    pfolder = ''
    for i in range(start_row, max_row+1):
        count+=1
        try:
            if(pfolder != ws.cell(row=i, column=2).value):
                pfolder = ws.cell(row=i, column=2).value
                os.makedirs(pfolder, exist_ok=True)
                #os.chdir(pfolder)
            fname = ws.cell(row=i, column=3).value
            open(pfolder+'\\'+fname, 'a').close()
            ws.cell(row=i, column=6, value='Prepared')
            success+=1
        except OSError as e:
            ws.cell(row=i, column=6, value = 'Error:Preparing')
            error+=1
            print('Error: ', e.args)
    report('Prepared dummy files for test', count, success, error)
    
def tst():
    global count, success, error
    os.chdir(tst_folder)
    fname_old = ''
    pfolder_old = ''
    fname_new = ''
    pfolder_new = ''
    to_move = True
    for i in range(start_row, max_row+1):
        count+=1
        try:
            pfolder_old = ws.cell(row=i, column=2).value
            fname_old = ws.cell(row=i, column=3).value
            pfolder_new = ws.cell(row=i, column=4).value
            if(pfolder_new == ''):
                pfolder_new = pfolder_old
                to_move = False
            fname_new = ws.cell(row=i, column=5).value
            if(fname_new == ''):
                fname_new = fname_old
                to_move = False
            #print('pfolder_old: ', pfolder_old, 'fname_old: ', fname_old, 'pfolder_new: ', pfolder_new, 'fname_new: ', fname_new, 'to_move: ', to_move,)
            if(to_move):
                to_move = True
                os.makedirs(pfolder_new, exist_ok=True)
                os.rename(pfolder_old+'\\'+fname_old, pfolder_new+'\\'+fname_new)
                ws.cell(row=i, column=6, value = 'Tested')
                success+=1
            else:
                ws.cell(row=i, column=6, value = 'No Changes')
                success+=1
        except OSError as e:
            ws.cell(row=i, column=6, value = 'Error:Test')
            error+=1
            print('Error: ', e.args)
    report('Test with dummy files', count, success, error)

def run():
    global count, success, error
    fname_old = ''
    pfolder_old = ''
    fname_new = ''
    pfolder_new = ''
    to_move = True
    for i in range(start_row, max_row+1):
        count+=1
        try:
            pfolder_old = ws.cell(row=i, column=2).value
            fname_old = ws.cell(row=i, column=3).value
            pfolder_new = ws.cell(row=i, column=4).value
            if(pfolder_new == ''):
                pfolder_new = pfolder_old
                to_move = False
            fname_new = ws.cell(row=i, column=5).value
            if(fname_new == ''):
                fname_new = fname_old
                to_move = False
            #print('pfolder_old: ', pfolder_old, 'fname_old: ', fname_old, 'pfolder_new: ', pfolder_new, 'fname_new: ', fname_new, 'to_move: ', to_move,)
            if(to_move):
                to_move = True
                os.makedirs(pfolder_new, exist_ok=True)
                os.rename(pfolder_old+'\\'+fname_old, pfolder_new+'\\'+fname_new)
                ws.cell(row=i, column=6, value = 'Moved')
                success+=1
            else:
                ws.cell(row=i, column=6, value = 'No Changes')
                success+=1
        except OSError as e:
            ws.cell(row=i, column=6, value = 'Error:Moving')
            error+=1
            print('Error: ', e.args)
    report('Files are renamed or moved', count, success, error)

def clean_fldrs(path):
    for root, fldrs, filenames in os.walk(path, topdown=False):
        for fldr in fldrs:
            print(os.path.realpath(os.path.join(root, fldr)))

def save_open():
    wb.save(sel_path+'\\'+filename)
    os.startfile(sel_path+'\\'+filename)

def report(action, count=0, success=0, error=0):
    print('\nAction performed: ', action)
    print('Total files handled: ', count)
    print('Success count: ', success)
    print('Failure count: ', error)

init()
